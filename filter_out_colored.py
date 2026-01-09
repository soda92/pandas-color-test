import copy
from openpyxl import load_workbook, Workbook

# --- CONFIGURATION ---
input_file = "qc_report_styled.xlsx"
output_file = "filtered_errors_colored.xlsx"

print(f"Loading {input_file}...")
wb_source = load_workbook(input_file) # Do NOT use data_only=True if you want to copy styles perfectly
ws_source = wb_source.active

# Create a new workbook for the output
wb_dest = Workbook()
ws_dest = wb_dest.active
ws_dest.title = "Filtered Errors"

def is_cell_colored(cell):
    """Checks if a cell has a non-white/non-transparent background."""
    fill = cell.fill
    if not fill.fill_type or fill.fill_type == 'none':
        return False
    
    # Check RGB (00000000 is transparent, FFFFFFFF is white)
    if fill.fgColor.type == 'rgb':
        if fill.fgColor.rgb in ['00000000', 'FFFFFFFF']:
            return False
        return True
    
    # Check Theme colors (usually implies a user selection)
    if fill.fgColor.type == 'theme':
        return True
        
    return False

def copy_row_style(source_row, dest_ws):
    """
    Appends a new row to dest_ws with values AND styles from source_row.
    """
    # 1. Append the values first to create the row
    values = [cell.value for cell in source_row]
    dest_ws.append(values)
    
    # 2. Get the newly created row in the destination
    #    (It will be the last row in the sheet)
    dest_row = dest_ws[dest_ws.max_row]
    
    # 3. Copy styles from source cell to destination cell
    for src_cell, dst_cell in zip(source_row, dest_row):
        # We must use copy() to avoid linking the objects
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

print("Processing rows...")

# 1. Always copy the header (Row 1)
header_row = ws_source[1]
copy_row_style(header_row, ws_dest)

# 2. Iterate through the rest of the data
#    (We assume headers are on row 1, data starts on row 2)
matched_count = 0

for row in ws_source.iter_rows(min_row=2):
    row_has_color = False
    
    # Check if ANY cell in this row is colored
    for cell in row:
        if is_cell_colored(cell):
            row_has_color = True
            break
    
    # If colored, copy the WHOLE row to the new file
    if row_has_color:
        copy_row_style(row, ws_dest)
        matched_count += 1

# 3. Save
wb_dest.save(output_file)
print(f"Done! Saved {matched_count} colored rows to '{output_file}'.")